"""
Facilitation Shuffle — Streamlit app for 1-2-4-all group generation.

Use this during sessions to:
  1. Check who is present (all ticked by default).
  2. Click Shuffle Groups to generate pairs → groups of 4 (round A) → groups of 4 (round B).
  3. Copy-paste the plain-text output into Zoom breakout rooms.
  4. Download the full summary as an xlsx to review all assignments at once.
"""
from __future__ import annotations

import io
import sys
import unicodedata
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "tmp" / "participants.xlsx"

# Make src/ importable so we can reuse the group-building logic.
sys.path.insert(0, str(ROOT))
from src.randomize_groups import _assign_ids, build_groups  # noqa: E402  (after sys.path patch)

# ---------------------------------------------------------------------------
# Helpers — load participants
# ---------------------------------------------------------------------------

def _load_participants_from_bytes(data: bytes) -> tuple[list[str], pd.DataFrame]:
    """
    Parse participant xlsx.

    Returns
    -------
    names : list[str]
        Sorted list of participant names (column A).
    df : pd.DataFrame
        All roster columns (A–D + present), sorted by name.
        Column E (present) is kept as-is from the file but overridden by
        the checkboxes in the UI.
    """
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    # Collect header row
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val).strip() if val else f"col_{col}")

    rows: list[dict] = []
    for row in range(2, ws.max_row + 1):
        name_val = ws.cell(row=row, column=1).value
        if name_val is None or not str(name_val).strip():
            continue
        row_dict: dict = {}
        for col_idx, header in enumerate(headers, start=1):
            row_dict[header] = ws.cell(row=row, column=col_idx).value
        rows.append(row_dict)

    wb.close()

    def _alpha_key(s: str) -> str:
        """Sort key: strip diacritics so Á/á sort with A, É/é with E, etc."""
        return "".join(
            c for c in unicodedata.normalize("NFD", s.lower())
            if unicodedata.category(c) != "Mn"
        )

    rows.sort(key=lambda r: _alpha_key(str(r.get(headers[0]) or "")))
    names = [str(r[headers[0]]).strip() for r in rows]

    # Keep only the first five columns (name, role, company, country, present)
    # so the DataFrame stays clean regardless of any pre-existing generated columns.
    keep = headers[:5]
    df = pd.DataFrame([{k: r.get(k) for k in keep} for r in rows])

    return names, df


@st.cache_data(show_spinner=False)
def _load_from_path(path: str) -> tuple[list[str], pd.DataFrame]:
    """Cached version for the default file path."""
    return _load_participants_from_bytes(Path(path).read_bytes())


# ---------------------------------------------------------------------------
# Helpers — group generation
# ---------------------------------------------------------------------------

def generate_groups(present: list[str]) -> dict[str, list[list[str]]]:
    """Run all three phases and return group assignments."""
    return build_groups(present)


def zoom_text(groups: list[list[str]], label: str = "Room") -> str:
    """Plain-text list ready to paste into Zoom breakout room names."""
    return "\n".join(f"{label} {i}: {', '.join(g)}" for i, g in enumerate(groups, 1))


def _build_summary_df(
    roster_df: pd.DataFrame,
    present_set: set[str],
    groups: dict[str, list[list[str]]],
) -> pd.DataFrame:
    """
    Combine roster with current group assignments into a single DataFrame.
    Non-present participants get empty group columns.
    """
    # Build lookup: name → group number
    name_col = roster_df.columns[0]

    pair_map = _assign_ids(groups["pairs"])
    g4a_map  = _assign_ids(groups["g4a"])
    g4b_map  = _assign_ids(groups["g4b"])

    df = roster_df.copy()
    df["present"]            = df[name_col].apply(lambda n: 1 if n in present_set else 0)
    df["pair (round 1)"]     = df[name_col].apply(lambda n: pair_map.get(n, pd.NA))
    df["group_4A (round 2)"] = df[name_col].apply(lambda n: g4a_map.get(n, pd.NA))
    df["group_4B (round 3)"] = df[name_col].apply(lambda n: g4b_map.get(n, pd.NA))

    # Cast to nullable integer so Arrow serialization works cleanly (absent rows → <NA>)
    for col in ["pair (round 1)", "group_4A (round 2)", "group_4B (round 3)"]:
        df[col] = df[col].astype(pd.Int64Dtype())

    return df


def _df_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    """Serialise a DataFrame to xlsx bytes with light formatting."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Groups")
        ws = writer.sheets["Groups"]
        # Bold headers + auto-width
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            width = max(len(str(c.value or "")) for c in col_cells) + 4
            ws.column_dimensions[col_cells[0].column_letter].width = min(width, 40)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Facilitation Shuffle",
    page_icon="🎲",
    layout="wide",
)

# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------
with st.sidebar:
    st.title("🎲 Facilitation Shuffle")
    st.caption("1-2-4-all group randomizer")
    st.divider()

    uploaded = st.file_uploader(
        "Upload participants list",
        type=["xlsx"],
        help="Leave empty to use **tmp/participants.xlsx** in the project folder.",
    )

    with st.expander("📋 Required file format"):
        st.markdown(
            """
**Excel (.xlsx) with at least column A:**

| Col | Header | Required? |
|-----|--------|-----------|
| A | `name` | ✅ Yes |
| B | `role` | optional |
| C | `company` | optional |
| D | `country` | optional |
| E | `present` | optional |

- **Row 1** must be a header row (any text).
- **Column A** must contain participant names — one per row, no blank names.
- **Column E** (`present`) is ignored by the app — use the checkboxes instead.
- Any pre-existing group columns (F–I) are ignored on upload.
"""
        )

    st.divider()
    st.caption("Generates: pairs → groups of 4 (A) → groups of 4 (B)")

# ---------------------------------------------------------------------------
# Load participant data
# ---------------------------------------------------------------------------
if uploaded is not None:
    raw_bytes = uploaded.read()
    all_names, roster_df = _load_participants_from_bytes(raw_bytes)
elif DEFAULT_XLSX.exists():
    all_names, roster_df = _load_from_path(str(DEFAULT_XLSX))
else:
    all_names, roster_df = [], pd.DataFrame()

# ---------------------------------------------------------------------------
# Main area
# ---------------------------------------------------------------------------
st.title("🎲 Facilitation Shuffle")

if not all_names:
    st.warning(
        "No participants found. "
        "Upload an .xlsx file using the sidebar, "
        "or place your file at `tmp/participants.xlsx`."
    )
    st.stop()

# ---------------------------------------------------------------------------
# Participant checklist
# ---------------------------------------------------------------------------
n_total = len(all_names)
st.subheader(f"Participants — {n_total} total")

col_sel, col_desel = st.columns([1, 1], gap="small")
with col_sel:
    if st.button("✅ Select all", width="stretch"):
        for i in range(n_total):
            st.session_state[f"p_{i}"] = True
with col_desel:
    if st.button("☐ Deselect all", width="stretch"):
        for i in range(n_total):
            st.session_state[f"p_{i}"] = False

# Initialise checkbox state on first load (all present by default).
# Must happen before rendering so that session_state owns the value —
# avoids the "created with default value but also set via Session State API" warning.
for i in range(n_total):
    if f"p_{i}" not in st.session_state:
        st.session_state[f"p_{i}"] = True

# Three-column checkbox grid — fill each column top-to-bottom so names read A-Z down each column
cols = st.columns(3)
checked: dict[str, bool] = {}
col_size = (n_total + 2) // 3  # ceiling division → equal chunks
for col_idx, col in enumerate(cols):
    with col:
        for i in range(col_idx * col_size, min((col_idx + 1) * col_size, n_total)):
            name = all_names[i]
            checked[name] = st.checkbox(name, key=f"p_{i}")

present = [name for name, ok in checked.items() if ok]
absent_count = n_total - len(present)
status_parts = [f"**{len(present)}** present"]
if absent_count:
    status_parts.append(f"**{absent_count}** absent")
st.caption(" · ".join(status_parts))

st.divider()

# ---------------------------------------------------------------------------
# Shuffle button
# ---------------------------------------------------------------------------
if st.button("🔀 Shuffle Groups", type="primary", width="stretch"):
    if len(present) < 2:
        st.error("Need at least 2 present participants to form groups.")
    else:
        _new_groups = generate_groups(present)
        st.session_state["groups"] = _new_groups
        st.session_state["present_set"] = set(present)
        st.session_state["n_present"] = len(present)
        st.session_state["ta_pairs"] = zoom_text(_new_groups["pairs"], "Room")
        st.session_state["ta_g4a"]   = zoom_text(_new_groups["g4a"],   "Room")
        st.session_state["ta_g4b"]   = zoom_text(_new_groups["g4b"],   "Room")

# ---------------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------------
if "groups" not in st.session_state:
    st.stop()

groups = st.session_state["groups"]
n_present = st.session_state["n_present"]
present_set: set[str] = st.session_state["present_set"]

st.subheader(f"Groups for {n_present} participants")

tab1, tab2, tab3, tab4 = st.tabs([
    "👥  Pairs (Round 1)",
    "👥👥  Groups of 4 — A (Round 2)",
    "👥👥  Groups of 4 — B (Round 3)",
    "📊  Summary",
])


def _render_round_tab(tab_id: str, label: str, phase_groups: list[list[str]], caption: str) -> None:
    st.caption(caption)

    st.text_area(
        "Copy and paste into Zoom:",
        value=zoom_text(phase_groups, label),
        height=max(120, len(phase_groups) * 30),
        key=f"ta_{tab_id}",
    )

    n_cols = min(len(phase_groups), 4)
    cards = st.columns(n_cols) if n_cols > 1 else [st.container()]
    for i, group in enumerate(phase_groups):
        with cards[i % n_cols]:
            members = "\n".join(f"- {p}" for p in group)
            st.markdown(f"**{label} {i + 1}**\n{members}")


with tab1:
    _render_round_tab(
        "pairs", "Room", groups["pairs"],
        "Use for the **pair reflection** phase — 2 people per room.",
    )

with tab2:
    _render_round_tab(
        "g4a", "Room", groups["g4a"],
        "Use for the **first group-of-4** phase (Personal Readme). "
        "Pairs from Round 1 are kept together.",
    )

with tab3:
    _render_round_tab(
        "g4b", "Room", groups["g4b"],
        "Use for the **second group-of-4** phase (Common Enemy). "
        "Optimised to maximise mixing across Round 2 groups.",
    )

with tab4:
    st.caption(
        "Full participant roster with all group assignments. "
        "Absent participants are shown with empty group columns."
    )

    summary_df = _build_summary_df(roster_df, present_set, groups)

    st.dataframe(
        summary_df,
        width="stretch",
        hide_index=True,
        column_config={
            "present":           st.column_config.NumberColumn("Present", width="small"),
            "pair (round 1)":    st.column_config.NumberColumn("Pair (R1)", width="small"),
            "group_4A (round 2)": st.column_config.NumberColumn("Group 4A (R2)", width="small"),
            "group_4B (round 3)": st.column_config.NumberColumn("Group 4B (R3)", width="small"),
        },
    )

    xlsx_bytes = _df_to_xlsx_bytes(summary_df)
    st.download_button(
        label="⬇️  Download as xlsx",
        data=xlsx_bytes,
        file_name="facilitation_groups.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
