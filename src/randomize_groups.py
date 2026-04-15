"""
Workshop group randomizer → Excel. See ``randomize_groups.md`` for full documentation.

Writes columns F–I from roster + ``present``; only rows with ``present == 1`` get group ids.
"""

from __future__ import annotations

import argparse
import random
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = PROJECT_ROOT / "tmp" / "participants.xlsx"

# Sheet layout: A–D = roster, E = present, F–I = generated group columns
COL_PRESENT = 5
COL_G2 = 6
COL_G4A = 7
COL_G4B = 8
COL_REPEAT = 9


def _is_present(value: object) -> bool:
    """Treat 1 /1.0 / \"1\" as present; everything else absent."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(value) == 1
    s = str(value).strip()
    return s == "1"


def _read_participants(ws) -> tuple[list[str], dict[str, int]]:
    """Return (names marked present, name → Excel row) for all rows with a name in column A."""
    present_names: list[str] = []
    row_map: dict[str, int] = {}
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1).value
        if cell is None or str(cell).strip() == "":
            continue
        name = str(cell).strip()
        row_map[name] = row
        if _is_present(ws.cell(row=row, column=COL_PRESENT).value):
            present_names.append(name)
    return present_names, row_map


def _all_pair_units(shuffled: list[str]) -> list[list[str]]:
    """Slice even-length list into consecutive pairs (phase-1 units of size 2)."""
    n = len(shuffled)
    if n % 2 != 0:
        raise ValueError(f"Expected even n for all-pairs phase; got {n}")
    return [shuffled[i : i + 2] for i in range(0, n, 2)]


def _trio_last_units(shuffled: list[str]) -> list[list[str]]:
    """For n ≡ 3 (mod 4): (n−3)/2 pairs, then one trio at the end — total n people."""
    n = len(shuffled)
    if n < 3 or n % 2 == 0:
        raise ValueError(f"Expected odd n >= 3 for trio-last; got {n}")
    units: list[list[str]] = []
    i = 0
    while i < n - 3:
        units.append(shuffled[i : i + 2])
        i += 2
    units.append(shuffled[n - 3 : n])
    return units


def _trio_first_units(shuffled: list[str]) -> list[list[str]]:
    """For n ≡ 1 (mod 4): one trio first, then pairs — keeps trio atomic for later merges."""
    n = len(shuffled)
    if n < 3 or n % 4 != 1:
        raise ValueError(f"Expected n ≡ 1 (mod 4) for trio-first; got n={n}")
    units: list[list[str]] = [shuffled[0:3]]
    i = 3
    while i < n:
        units.append(shuffled[i : i + 2])
        i += 2
    return units


def _merge_adjacent_units(pair_units: list[list[str]]) -> list[list[str]]:
    """
    Phase-2 from ordered units: concatenate unit[0]+unit[1], unit[2]+unit[3], …
    If odd count of units, the last unit stands alone (e.g. a trio after pairs).
    Never splits a unit — couples/trios stay together.
    """
    groups: list[list[str]] = []
    p = len(pair_units)
    j = 0
    while j < p:
        if j + 1 < p:
            groups.append(pair_units[j] + pair_units[j + 1])
            j += 2
        else:
            groups.append(pair_units[j])
            j += 1
    return groups


def _merge_pair_units_mod4_eq2(pair_units: list[list[str]]) -> list[list[str]]:
    """
    n ≡ 2 (mod 4), all units are pairs: (2k+1) pair-units.

    First (k−1) merges are 2+2 → 4; the last merge is three pairs → 6 (no leftover pair of 2).
    n == 2 → single pair only.
    """
    p = len(pair_units)
    if p == 1:
        return [pair_units[0]]
    k = (p - 1) // 2
    groups: list[list[str]] = []
    idx = 0
    for _ in range(k - 1):
        groups.append(pair_units[idx] + pair_units[idx + 1])
        idx += 2
    groups.append(pair_units[idx] + pair_units[idx + 1] + pair_units[idx + 2])
    return groups


def _build_phase1_units(shuffled: list[str], n: int) -> list[list[str]]:
    """
    Build phase-1 units (pair/trio blocks) from a shuffled name list.

    Pattern by n mod 4: all pairs; pairs+trio last; all pairs (n mod 4 == 2); trio first + pairs (n mod 4 == 1).
    """
    if n == 1:
        return [shuffled[0:1]]
    r = n % 4
    if r == 0:
        return _all_pair_units(shuffled)
    if r == 3:
        return _trio_last_units(shuffled)
    if r == 2:
        return _all_pair_units(shuffled)
    return _trio_first_units(shuffled)


def _build_phase2_groups(pair_units: list[list[str]], n: int) -> list[list[str]]:
    """
    Phase-2 = merge whole phase-1 units only (never shuffle people across units).

    If n mod 4 == 2: blocks of 4 plus one block of 6; else adjacent 2+2 merges (or lone last unit).
    """
    r = n % 4
    if r == 2:
        return _merge_pair_units_mod4_eq2(pair_units)
    return _merge_adjacent_units(pair_units)


def _phase3_target_sizes(n: int) -> list[int]:
    """
    Sizes for phase-3 groups: prefer 4, then 3; leftovers absorbed into last group.
    Guarantees each group ≥ 3 when n ≥ 3 (except tiny n).
    """
    if n <= 0:
        return []
    if n < 3:
        return [n]
    sizes: list[int] = []
    rem = n
    while rem > 0:
        # Take 4 only if remainder stays 0 or ≥3 (avoid stranding 1–2 alone)
        if rem >= 4 and (rem - 4 == 0 or rem - 4 >= 3):
            sizes.append(4)
            rem -= 4
        elif rem >= 3:
            sizes.append(3)
            rem -= 3
        else:
            if sizes:
                sizes[-1] += rem
            else:
                sizes.append(rem)
            rem = 0
    return sizes


def _slice_groups_in_order(order: list[str], sizes: list[int]) -> list[list[str]]:
    """Cut a permutation into contiguous chunks of given sizes (phase-3 trial layout)."""
    groups: list[list[str]] = []
    k = 0
    for s in sizes:
        groups.append(order[k : k + s])
        k += s
    return groups


def _phase2_cohort_overlap_penalty(
    groups_phase3: list[list[str]], person_to_phase2_idx: dict[str, int]
) -> int:
    """
    Lower is better: for each phase-3 group, count (c−1) when c>1 people share the same
    phase-2 table id (extra “repeats” vs ideal mixing).
    """
    pen = 0
    for g in groups_phase3:
        counts: dict[int, int] = {}
        for p in g:
            i = person_to_phase2_idx[p]
            counts[i] = counts.get(i, 0) + 1
        for c in counts.values():
            if c > 1:
                pen += c - 1
    return pen


def _build_phase3_relaxed(
    shuffled: list[str],
    n: int,
    person_to_phase2_idx: dict[str, int],
    iterations: int = 120_000,
) -> tuple[list[list[str]], int]:
    """
    Random search: shuffle names, slice by _phase3_target_sizes, keep layout with lowest
    phase-2 overlap penalty. Allows same phase-2 table twice in one phase-3 group if needed.
    """
    sizes = _phase3_target_sizes(n)
    assert sum(sizes) == n, (sizes, n)
    pool = shuffled[:]
    best: list[list[str]] | None = None
    best_pen = sys.maxsize
    for _ in range(iterations):
        random.shuffle(pool)
        trial = _slice_groups_in_order(pool, sizes)
        pen = _phase2_cohort_overlap_penalty(trial, person_to_phase2_idx)
        if pen < best_pen:
            best_pen = pen
            best = [x[:] for x in trial]
        if best_pen == 0:
            break
    assert best is not None
    return best, best_pen


def _assign_ids(groups: list[list[str]]) -> dict[str, int]:
    """Map each person to 1-based group index within that phase."""
    out: dict[str, int] = {}
    for idx, g in enumerate(groups, start=1):
        for p in g:
            out[p] = idx
    return out


def run(path: Path, seed: int | None = None) -> None:
    if seed is not None:
        random.seed(seed)

    wb = load_workbook(path)
    ws = wb.active

    present_names, row_map = _read_participants(ws)
    n = len(present_names)
    print(f"Present (count): {n}")

    all_names = list(row_map.keys())

    if n == 0:
        print("No participants with present=1; clearing group columns only.")
        id_pair: dict[str, int] = {}
        id_f: dict[str, int] = {}
        id_g: dict[str, int] = {}
        person_to_f_set: dict[str, set[str]] = {}
        person_to_g_set: dict[str, set[str]] = {}
    else:
        shuffled = present_names[:]
        random.shuffle(shuffled)

        pair_units = _build_phase1_units(shuffled, n)
        print(f"Phase1 units: {len(pair_units)} -> sizes {[len(u) for u in pair_units]}")

        groups_f = _build_phase2_groups(pair_units, n)
        print(f"Phase2 (personal_readme_g4): {len(groups_f)} -> sizes {[len(g) for g in groups_f]}")

        person_to_phase2_idx: dict[str, int] = {}
        for idx, g in enumerate(groups_f):
            for p in g:
                person_to_phase2_idx[p] = idx

        groups_g, p3_pen = _build_phase3_relaxed(shuffled, n, person_to_phase2_idx)
        sz = [len(g) for g in groups_g]
        print(
            f"Phase3 (common_enemy_g4): {len(groups_g)} -> sizes {sz} "
            f"(min={min(sz)}; phase2-cohort overlap penalty={p3_pen})"
        )

        id_pair = _assign_ids(pair_units)
        id_f = _assign_ids(groups_f)
        id_g = _assign_ids(groups_g)

        person_to_f_set = {p: set(g) for g in groups_f for p in g}
        person_to_g_set = {p: set(g) for g in groups_g for p in g}

    # Headers for generated columns only (E = present is user-owned)
    ws.cell(row=1, column=COL_G2).value = "personal_readme_g2"
    ws.cell(row=1, column=COL_G4A).value = "personal_readme_g4"
    ws.cell(row=1, column=COL_G4B).value = "common_enemy_g4"
    ws.cell(row=1, column=COL_REPEAT).value = "group_repeat"
    for col in (COL_G2, COL_G4A, COL_G4B, COL_REPEAT):
        c = ws.cell(row=1, column=col)
        c.font = Font(name="Arial", bold=True)
        c.alignment = Alignment(horizontal="center")

    present_set = set(present_names)

    for name in all_names:
        row = row_map[name]
        if name not in present_set:
            for col in (COL_G2, COL_G4A, COL_G4B, COL_REPEAT):
                ws.cell(row=row, column=col).value = None
            continue

        if n == 0:
            continue

        ws.cell(row=row, column=COL_G2).value = id_pair[name]
        ws.cell(row=row, column=COL_G4A).value = id_f[name]
        ws.cell(row=row, column=COL_G4B).value = id_g[name]

        # Sheet column “group_repeat”: other people shared with in both phase 2 and phase 3
        f_others = person_to_f_set[name] - {name}
        g_others = person_to_g_set[name] - {name}
        ws.cell(row=row, column=COL_REPEAT).value = len(f_others & g_others)

        for col in (COL_G2, COL_G4A, COL_G4B, COL_REPEAT):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(horizontal="center")

    for letter in ("F", "G", "H", "I"):
        ws.column_dimensions[letter].width = 14

    wb.save(path)
    print(f"Saved: {path}")


def _make_fixture_xlsx(path: Path, n: int = 15) -> None:
    """Minimal workbook for testing: name + present=1."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    headers = ["name", "role", "company", "country", "present"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = h
    for i in range(n):
        r = 2 + i
        ws.cell(row=r, column=1).value = f"Participant {i + 1}"
        ws.cell(row=r, column=COL_PRESENT).value = 1
    wb.save(path)
    print(f"Wrote fixture: {path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Randomize pair and group columns in participant xlsx.")
    parser.add_argument(
        "xlsx",
        nargs="?",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"Workbook path (default: {DEFAULT_XLSX})",
    )
    parser.add_argument("--seed", type=int, default=None, help="RNG seed for reproducible runs")
    parser.add_argument(
        "--make-fixture",
        action="store_true",
        help=f"Create {DEFAULT_XLSX} with 15 placeholder names if missing, then exit",
    )
    args = parser.parse_args()

    if args.make_fixture and not args.xlsx.exists():
        _make_fixture_xlsx(args.xlsx)
        return

    if not args.xlsx.exists():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        print("Run with --make-fixture to create a sample workbook.", file=sys.stderr)
        sys.exit(1)

    run(args.xlsx, seed=args.seed)


if __name__ == "__main__":
    main()
