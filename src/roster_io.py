"""
Roster loading — parse a participant xlsx into names + a clean DataFrame.

Pure data logic (pandas/openpyxl only); never imports streamlit.
"""
from __future__ import annotations

import io
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def _alpha_key(s: str) -> str:
    """Sort key: strip diacritics so Á/á sort with A, É/é with E, etc."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower())
        if unicodedata.category(c) != "Mn"
    )


def load_participants_from_bytes(data: bytes) -> tuple[list[str], pd.DataFrame]:
    """
    Parse participant xlsx.

    Returns
    -------
    names : list[str]
        Sorted list of participant names (column A).
    df : pd.DataFrame
        All roster columns (A–D + present), sorted by name.
        Column E (present) is kept as-is from the file but overridden by
        the checkboxes in the UI.
    """
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    # Collect header row
    headers: list[str] = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val).strip() if val else f"col_{col}")

    rows: list[dict] = []
    for row in range(2, ws.max_row + 1):
        name_val = ws.cell(row=row, column=1).value
        if name_val is None or not str(name_val).strip():
            continue
        row_dict: dict = {}
        for col_idx, header in enumerate(headers, start=1):
            row_dict[header] = ws.cell(row=row, column=col_idx).value
        rows.append(row_dict)

    wb.close()

    rows.sort(key=lambda r: _alpha_key(str(r.get(headers[0]) or "")))
    names = [str(r[headers[0]]).strip() for r in rows]

    # Keep only the first five columns (name, role, company, country, present)
    # so the DataFrame stays clean regardless of any pre-existing generated columns.
    keep = headers[:5]
    df = pd.DataFrame([{k: r.get(k) for k in keep} for r in rows])

    return names, df


def load_participants_from_path(path: str) -> tuple[list[str], pd.DataFrame]:
    """Load participants from a file path on disk."""
    return load_participants_from_bytes(Path(path).read_bytes())
